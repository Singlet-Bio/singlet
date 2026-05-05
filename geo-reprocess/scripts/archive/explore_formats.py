"""Hands-on exploration of 2c candidate formats - download & inspect real files."""
import pandas as pd
import numpy as np
import os, sys, re, json, tempfile, tarfile, io, struct
import urllib.request
import gzip

TMPDIR = "/tmp/2c_explore"
os.makedirs(TMPDIR, exist_ok=True)

# --- 1. EXPLORE TAR FILES ---
print("=" * 80)
print("1. TAR FILES (15,140 occurrences)")
print("=" * 80)

# Download just the first few KB of a RAW.tar to see the directory listing
tar_url = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE100nnn/GSE100274/suppl/GSE100274_RAW.tar"
print(f"\nFetching TOC from: {tar_url}")
try:
    req = urllib.request.Request(tar_url)
    # Download first 1MB to get the file listing
    req.add_header('Range', 'bytes=0-1048576')
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    # Parse as tar
    tarbuf = io.BytesIO(data)
    with tarfile.open(fileobj=tarbuf, mode='r|') as tf:
        members = []
        for i, m in enumerate(tf):
            if i > 30: break
            members.append(m.name)
    print(f"  First {len(members)} entries:")
    for m in members[:20]:
        print(f"    {m}")
except Exception as e:
    print(f"  Error: {e}")

# Try another GSE's tar
print()
tar_url2 = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE116nnn/GSE116390/suppl/GSE116390_VDJ_annotations.csv.tar.gz"
print(f"Fetching: {os.path.basename(tar_url2)}")
try:
    outfile = os.path.join(TMPDIR, "vdj_annot.tar.gz")
    urllib.request.urlretrieve(tar_url2, outfile)
    with tarfile.open(outfile, 'r:gz') as tf:
        members = tf.getnames()[:10]
        print(f"  {len(tf.getnames())} files in archive")
        print(f"  First entries: {members}")
        # Extract first file and peek
        for m in tf:
            if m.isfile() and m.size < 10_000_000:
                f = tf.extractfile(m)
                if f:
                    content = f.read(2048)
                    if m.name.endswith('.gz'):
                        content = gzip.decompress(content[:min(len(content), 50000)])[:2048]
                    lines = content.decode('utf-8', errors='replace').split('\n')[:5]
                    print(f"  Peek at {m.name}:")
                    for line in lines:
                        print(f"    {line[:120]}")
                    break
except Exception as e:
    print(f"  Error: {e}")

# --- 2. EXPLORE H5 FILES (10x CellRanger h5 format) ---
print("\n" + "=" * 80)
print("2. H5 FILES (6,534 occurrences)")
print("=" * 80)

import h5py

h5_url = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE117nnn/GSE117963/suppl/GSE117963_10X_lineage_positive_filtered_gene_bc_matrices_h5.h5"
print(f"\nDownloading: {os.path.basename(h5_url)}")
try:
    outfile = os.path.join(TMPDIR, "test.h5")
    urllib.request.urlretrieve(h5_url, outfile)
    with h5py.File(outfile, 'r') as f:
        def print_tree(g, prefix=''):
            for key in list(g.keys())[:15]:
                item = g[key]
                if isinstance(item, h5py.Group):
                    print(f"  {prefix}{key}/ ({len(item)} items)")
                    print_tree(item, prefix + '  ')
                elif isinstance(item, h5py.Dataset):
                    print(f"  {prefix}{key} {item.shape} {item.dtype}")
        print("  H5 tree:")
        print_tree(f)
        # Check for barcodes
        for path in ['matrix/barcodes', 'barcodes', '/matrix/barcodes']:
            if path in f:
                barcodes = f[path][:5]
                print(f"\n  Barcodes at '{path}' (first 5): {[b.decode() if isinstance(b, bytes) else b for b in barcodes]}")
                break
except Exception as e:
    print(f"  Error: {e}")

# --- 3. EXPLORE XLSX FILES ---
print("\n" + "=" * 80)
print("3. XLSX FILES (328 occurrences)")
print("=" * 80)

xlsx_url = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE109nnn/GSE109447/suppl/GSE109447_13055_Cell_Barcodes.xlsx"
print(f"\nDownloading: {os.path.basename(xlsx_url)}")
try:
    outfile = os.path.join(TMPDIR, "test.xlsx")
    urllib.request.urlretrieve(xlsx_url, outfile)
    import openpyxl
    wb = openpyxl.load_workbook(outfile, read_only=True)
    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        print(f"\n  Sheet '{sheet_name}':")
        rows = list(ws.iter_rows(max_row=6, values_only=True))
        for row in rows[:6]:
            print(f"    {[str(c)[:30] if c else '' for c in row[:10]]}")
    wb.close()
except ImportError:
    print("  openpyxl not installed - trying pandas")
    try:
        df = pd.read_excel(outfile)
        print(f"  Shape: {df.shape}")
        print(f"  Columns: {list(df.columns[:15])}")
        print(f"  Head:\n{df.head()}")
    except Exception as e:
        print(f"  Error: {e}")
except Exception as e:
    print(f"  Error: {e}")

# --- 4. EXPLORE H5SEURAT ---
print("\n" + "=" * 80)
print("4. H5SEURAT FILES (24 occurrences)")  
print("=" * 80)

h5s_url = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE178nnn/GSE178088/suppl/GSE178088_cluster3.h5Seurat"
print(f"\nDownloading: {os.path.basename(h5s_url)}")
try:
    outfile = os.path.join(TMPDIR, "test.h5seurat")
    urllib.request.urlretrieve(h5s_url, outfile)
    with h5py.File(outfile, 'r') as f:
        def print_tree2(g, prefix='', depth=0):
            if depth > 3: return
            for key in list(g.keys())[:20]:
                item = g[key]
                if isinstance(item, h5py.Group):
                    print(f"  {prefix}{key}/ ({len(item)} items)")
                    print_tree2(item, prefix + '  ', depth+1)
                elif isinstance(item, h5py.Dataset):
                    s = str(item.shape)
                    print(f"  {prefix}{key} {s} {item.dtype}")
        print("  H5Seurat tree:")
        print_tree2(f)
        # Try to find cell metadata
        for path in ['meta.data', 'cell.names', 'meta.features']:
            if path in f:
                grp = f[path]
                if isinstance(grp, h5py.Group):
                    print(f"\n  {path}/ keys: {list(grp.keys())[:20]}")
                    for k in list(grp.keys())[:3]:
                        d = grp[k]
                        if isinstance(d, h5py.Dataset):
                            vals = d[:5]
                            print(f"    {k}: {[v.decode() if isinstance(v, bytes) else v for v in vals]}")
                elif isinstance(grp, h5py.Dataset):
                    print(f"\n  {path}: {grp[:5]}")
except Exception as e:
    print(f"  Error: {e}")

# --- 5. EXPLORE RDATA ---
print("\n" + "=" * 80)
print("5. RDATA/RDA FILES (210 occurrences)")
print("=" * 80)
print("  RData files contain R workspaces (multiple objects)")
print("  Requires rpy2 + R to load, similar to RDS but can have multiple objects")
print("  Will treat same as stage2b extension")

# --- 6. EXPLORE HDF5 ---
print("\n" + "=" * 80)
print("6. GENERIC HDF5 FILES (28 occurrences)")
print("=" * 80)

hdf5_url = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE245nnn/GSE245419/suppl/GSE245419.hdf5"
print(f"\nDownloading: {os.path.basename(hdf5_url)} (may be large)")
try:
    outfile = os.path.join(TMPDIR, "test.hdf5")
    # Only download first 10MB
    req = urllib.request.Request(hdf5_url)
    req.add_header('Range', 'bytes=0-10485760')
    with urllib.request.urlopen(req, timeout=60) as resp:
        with open(outfile, 'wb') as fout:
            fout.write(resp.read())
    with h5py.File(outfile, 'r') as f:
        print("  HDF5 tree:")
        print_tree2(f)
except Exception as e:
    print(f"  Error (expected for partial download): {e}")

# --- 7. EXPLORE H5MU (MuData) ---
print("\n" + "=" * 80)
print("7. H5MU / MUDATA (0 in processed, 3424 in full catalog)")
print("=" * 80)
print("  MuData format has .obs (cell metadata) similar to h5ad")
print("  Can read with mudata package or h5py directly")
print("  obs/ group contains per-cell annotations")

# --- SUMMARY ---
print("\n" + "=" * 80)
print("SUMMARY: Stage 2c Implementation Plan")
print("=" * 80)
print("""
Priority formats for 2c (by processed GSM count):
  1. TAR archives     (15,140) - Extract per-GSM metadata files from RAW.tar
  2. H5/CellRanger    (6,534)  - Read barcodes (expression only, no metadata cols)
  3. XLSX/XLS         (332)    - Cell barcodes, annotations in Excel
  4. RData/RDA/RObj   (216)    - R workspace objects → treat like RDS
  5. H5Seurat         (24)     - HDF5-based Seurat → read meta.data/ group
  6. HDF5 generic     (28)     - Inspect structure, extract if obs-like
  
Key insight: TAR files are the biggest win. These are GSE_RAW.tar archives
that contain per-GSM supplementary files. Many of these are 10x barcodes/matrix
bundles, but some contain metadata files specific to each sample.
The current pipeline only looks at GSE-level supplementary files listed in the catalog.
The TAR files contain per-GSM files that we've been missing entirely.
""")
